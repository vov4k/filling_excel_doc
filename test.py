import openpyxl as opx
import shutil
import sys  # sys нужен для передачи argv в QApplication
from PySide2 import QtWidgets
import design  # Это наш конвертированный файл дизайна

class ExampleApp(QtWidgets.QMainWindow, design.Ui_MainWindow):
    dir = ''
    settings = ''
    db = ''
    template = ''
    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна
        self.save_dir.clicked.connect(self.browse_folder)
        self.settings_file.clicked.connect(self.browse_settings)
        self.data_base.clicked.connect(self.browse_db)
        self.tmplate_file.clicked.connect(self.browse_template)
        self.Start.clicked.connect(self.start_proc)

    def browse_folder(self):
        self.dir = QtWidgets.QFileDialog.getExistingDirectory(self, "Выберите папку")
        if self.dir:
              self.save_dir.setText("Выбрано")
    def browse_settings(self):
        self.settings = QtWidgets.QFileDialog.getOpenFileName(self, "Выберите файл", "", "Файл настроек (*.template)")[0]
        if self.settings:
              self.settings_file.setText("Выбрано")
    def browse_db(self):
        self.db = QtWidgets.QFileDialog.getOpenFileName(self, "Выберите файл", "", "Excel файл (*.xlsx)")[0]
        if self.db:
              self.data_base.setText("Выбрано")
    def browse_template(self):
        self.template = QtWidgets.QFileDialog.getOpenFileName(self, "Выберите файл", "", "Файл шаблона (*.xlsx *.docx)")[0]
        if self.template:
              self.tmplate_file.setText("Выбрано")
    def start_proc(self):
            settings_file = self.settings
            file_with_data = self.db
            template = self.template
            fio_wb = opx.load_workbook(file_with_data)
            fio_sheet = fio_wb.worksheets[0]
            n_row = 0
            
            with open(settings_file) as file:
                indexes = list()
                for line in file.readlines(): 
                    indexes.append(line.rstrip().split(';')) # Обрезаем каретку переноса строки и делим по пробелу
            
            print("Введите (через пробел) номера полей которые будут в названии файла")
            k = 1
            for temp in indexes:
                    print(k, "-", temp[0])
                    k+=1
            names = list(map(int, input("\n").split()))
            
            for row in fio_sheet.rows:
                    data_one_row = []
                    
                    for d in row:
                            data_one_row.append(str(d.value))
                    if n_row == 0:
                            n_row = 1
                            headers=list(data_one_row)
                            print(headers)
                            continue
                    
                    
                    file = self.dir+"/"
                    for h in names:
                            if data_one_row[h-1]:
                                    file+=data_one_row[h-1].strip()+" "
                    
                    file =file.strip() + ".xlsx"
                    print("Начал: "+file)
                    shutil.copy(template, file)
                    
                    data_wb = opx.load_workbook(file)
                    sheet = data_wb.worksheets[0]
                    j = 0
            
                    lll=0
                    for tmp in indexes:
                    
                        for header in headers:
                            if header == tmp[0]:
                                one_data=data_one_row[lll]
                                if one_data and one_data != 'None':
                                    ji = 0
                                    for i in range(len(one_data)):
                                            ffff = sheet.cell(row=int(indexes[j][1]), column=ji+int(indexes[j][2])).coordinate
                                            sheet[ffff].value=one_data[i]
                                            ji+=int(indexes[j][3])
                                j = j +1
                                lll+=1
                                break
                    data_wb.save(file)
                    print("Закончил: "+file)
            self.progressBar.setValue(100)

def main():
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение

if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()

