import shutil
import sys  # sys нужен для передачи argv в QApplication
import time

import openpyxl as opx
from PySide2 import QtWidgets
from PySide2.QtWidgets import QDialog

import design  # Это наш конвертированный файл дизайна
import form_template_name


class choose_file_name_win(QDialog, form_template_name.Ui_name_template):
    def __init__(self, name_templates, mainwindow):
        QDialog.__init__(self)
        self.mainwindow=mainwindow
        self.setupUi(self)

        self.mainwindow.flag = False
        self.listWidget.addItems(name_templates)
        self.listWidget.itemDoubleClicked.connect(self.add_to_input)
        self.buttonBox.accepted.connect(self.acept_data)
        self.buttonBox.rejected.connect(self.reject_data)
    def add_to_input(self, item_obj):
        text = item_obj.text().split(" - ")
        text = "#"+str(text[0])+"%"
        old_text = self.template_of_file_name.text()
        self.template_of_file_name.setText(old_text+text)
    def acept_data(self):
        self.mainwindow.file_name_template=self.template_of_file_name.text()
        self.close()
    def reject_data(self):
        self.close()


class ExampleApp(QtWidgets.QMainWindow, design.Ui_MainWindow):
    dir = ''
    settings = ''
    db = ''
    flag = True
    template = ''
    file_name_template = ''
    templates_names = ''
    indexes = list()
    def __init__(self):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна

        self.save_dir.clicked.connect(self.browse_folder)
        self.settings_file.clicked.connect(self.browse_settings)
        self.data_base.clicked.connect(self.browse_db)
        self.tmplate_file.clicked.connect(self.browse_template)
        self.Start.clicked.connect(self.choose_template)

    def browse_folder(self):
        self.dir = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Выберите папку")
        if self.dir:
            self.save_dir.setText("Выбрано")

    def browse_settings(self):
        self.settings = QtWidgets.QFileDialog.getOpenFileName(
            self, "Выберите файл настроек", "", "Файл настроек (*.template)")[0]
        if self.settings:
            self.settings_file.setText("Выбрано")

    def browse_db(self):
        tmp = QtWidgets.QFileDialog.getOpenFileName(
            self, "Выберите файл с БД", "", "Excel файл (*.xlsx)")[0]
        if tmp:
            self.db=tmp
            self.data_base.setText("Выбрано")

    def browse_template(self):
        self.template = QtWidgets.QFileDialog.getOpenFileName(
            self, "Выберите шаблон", "", "Файл шаблона (*.xlsx *.docx)")[0]
        if self.template:
            self.tmplate_file.setText("Выбрано")
    def main_win(self):
        file_with_data = self.db
        template = self.template
        fio_wb = opx.load_workbook(file_with_data, data_only=True, read_only=True)
        fio_sheet = fio_wb.worksheets[0]
        n_row = 0
        max_row = fio_sheet.max_row
        for row in fio_sheet.rows:
            data_one_row = []
            for d in row:
                data_one_row.append(str(d.value))
            if n_row == 0:
                n_row = 1
                headers = list(data_one_row)
                continue
            file = self.dir+"/" 
            file_name_template_copy=str(self.file_name_template)
            for ggg in range(len(data_one_row)):
                if str(ggg+1) in file_name_template_copy:
                    file_name_template_copy = file_name_template_copy.replace(str(ggg+1),data_one_row[ggg])
            file_name_template_copy = file_name_template_copy.replace("#","")
            file_name_template_copy = file_name_template_copy.replace("%","")
            file += file_name_template_copy
            file = file.strip() + ".xlsx"
            print("Начал: "+file)
            shutil.copy(template, file)
            data_wb = opx.load_workbook(file)
            sheet = data_wb.worksheets[0]
            j = 0
            for tmp in self.indexes:
                lll = 0  # номер столбца с данными
                for header in headers:
                    if header == tmp[0]:
                        one_data = data_one_row[lll]
                        if one_data and one_data != 'None':
                            ji = 0
                            for i in range(len(one_data)):
                                ffff = sheet.cell(
                                    row=int(self.indexes[j][1]), column=ji+int(self.indexes[j][2])).coordinate
                                sheet[ffff].value = one_data[i]
                                ji += int(self.indexes[j][3])
                        j = j + 1
                        break
                    lll += 1
            data_wb.save(file)
            print("Закончил: "+file)
            self.progressBar.setValue((n_row/max_row)*100)
            n_row += 1
        self.progressBar.setValue(100)

    def choose_template(self):
        if self.flag:
            with open(self.settings) as file:
                for line in file.readlines():
                    # Обрезаем каретку переноса строки и делим по пробелу
                    self.indexes.append(line.rstrip().split(';'))
            self.templates_names=[]
            k = 1
            for temp in self.indexes:
                self.templates_names.append(str(k)+" - "+temp[0])
                k += 1
        self.dialog = choose_file_name_win(self.templates_names,self)
        self.dialog.show()
        self.dialog.exec_()
        if self.file_name_template:
            self.main_win()


def main():
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение


if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()
